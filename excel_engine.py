import openpyxl
import shutil
import os
from datetime import datetime

# --- CONFIGURAÇÃO (TAGS) ---
TAGS_FIXAS = {
    "Salgados.quentes": "{list_salg_qts}",
    "Salgados.frios": "{list_salg_fr}",
    "Salgados.assados": "{list_salg_ass}",
    "Salgados.petit_gourmet": "{list_salg_pg}",
    "Bebidas": "{list_drinks}",
    "Sobremesa": "{list_dessert}", 
    "Buffet Infantil": "{list_infantil}",
    "Mão de Obra": "{list_hw}" # Tag da Mão de Obra
}

TAGS_SIMPLES = {
    "cliente": "{name_client}",
    "data": "{date}",
    "convidados": "{num_guest}", 
    "local": "{local}",
    "valor_total": "{valor_total}", # Tag do Valor
    "tipo": "{type_recepcao}" # Tag do Tipo (Cabeçalho)
}

# --- HELPER FUNCTIONS ---
def buscar_dados_por_caminho(dados_completos, caminho_string):
    chaves = caminho_string.split(".")
    conteudo = dados_completos.get('cardapio', {})
    try:
        for chave in chaves:
            if isinstance(conteudo, dict):
                conteudo = conteudo.get(chave, [])
            else: return []
        return conteudo if isinstance(conteudo, list) else []
    except: return []

def substituir_tag_simples(ws, tag, valor):
    """Substitui tag por valor mesmo que haja texto ao redor"""
    val_str = str(valor)
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if tag in cell.value:
                    cell.value = cell.value.replace(tag, val_str)

def expandir_lista_no_excel(ws, tag_excel, lista_dados):
    linha_tag = -1
    coluna_tag = -1
    
    # 1. Acha Tag
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == tag_excel:
                linha_tag = cell.row
                coluna_tag = cell.column
                break
        if linha_tag != -1: break
    
    if linha_tag == -1: return False
    
    # 2. Preenche ou Deleta
    if not lista_dados:
        try:
            ws.delete_rows(linha_tag - 1, amount=3)
        except: pass
    else:
        qtd = len(lista_dados)
        ws.cell(row=linha_tag, column=coluna_tag).value = lista_dados[0]
        if qtd > 1:
            ws.insert_rows(linha_tag + 1, amount=qtd - 1)
            for i in range(1, qtd):
                ws.cell(row=linha_tag + i, column=coluna_tag).value = lista_dados[i]
    return True

# --- FUNÇÃO PRINCIPAL ---
def gerar_excel(dados_json):
    pasta_data = "data"
    arquivo_template = os.path.join(pasta_data, "template_orcamento.xlsx")
    
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    cliente_safe = str(dados_json['metadados']['cliente'] or "SemNome").replace(' ', '_')
    nome_saida = f"Orcamento_{cliente_safe}_{ts}.xlsx"
    caminho_saida = os.path.join("output", nome_saida)

    if not os.path.exists("output"): os.makedirs("output")
    try: shutil.copy(arquivo_template, caminho_saida)
    except: return {"sucesso": False, "erro": "Template não encontrado!"}

    wb = openpyxl.load_workbook(caminho_saida)
    ws = wb.active

    # 1. TAGS SIMPLES (Valor, Cliente, Data, etc.)
    for chave, tag in TAGS_SIMPLES.items():
        val = dados_json['metadados'].get(chave, "")
        substituir_tag_simples(ws, tag, val)
    
    substituir_tag_simples(ws, "{obs}", dados_json.get("obs", ""))

    # 2. LISTAS PADRÃO (Salgados, Mão de Obra, etc.)
    for caminho, tag in TAGS_FIXAS.items():
        lista = buscar_dados_por_caminho(dados_json, caminho)
        expandir_lista_no_excel(ws, tag, lista)

    # 3. LISTA ESPECÍFICA DO TIPO (Boteco, Infantil, etc.)
    tipo = dados_json['metadados'].get('tipo')
    lista_tipo = dados_json['cardapio'].get(tipo, [])
    # Fallback se não achar lista com nome exato do tipo
    if not lista_tipo and tipo == "Infantil": 
        lista_tipo = dados_json['cardapio'].get("Buffet Infantil", [])
    
    expandir_lista_no_excel(ws, "{list_type_recp}", lista_tipo)

    # 4. PRATO PRINCIPAL (Títulos Dinâmicos)
    dados_prato = dados_json['cardapio'].get('Prato Principal', {})
    cats = [(k, v) for k, v in dados_prato.items() if v]

    # Slot 1
    if len(cats) > 0:
        nome_1, lista_1 = cats[0]
        substituir_tag_simples(ws, "{1_opcao}", nome_1.upper())
        expandir_lista_no_excel(ws, "{list_1_op}", lista_1)
    else:
        expandir_lista_no_excel(ws, "{list_1_op}", [])

    # Slot 2
    if len(cats) > 1:
        nome_2, lista_2 = cats[1]
        substituir_tag_simples(ws, "{2_opcao}", nome_2.upper())
        expandir_lista_no_excel(ws, "{list_2_op}", lista_2)
    else:
        substituir_tag_simples(ws, "{2_opcao}", "")
        expandir_lista_no_excel(ws, "{list_2_op}", [])

    wb.save(caminho_saida)
    return {"sucesso": True, "caminho": caminho_saida}