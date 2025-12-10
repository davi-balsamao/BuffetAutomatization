import streamlit as st
import openpyxl
import os

st.title("🚀 Painel de Controle: Buffet Automatization")
st.write("Se você está vendo isso, o Streamlit está rodando no WSL!")

st.header("Teste de Criação de Arquivo")
nome_arquivo = st.text_input("Nome do arquivo (sem extensão):", value="teste_wsl")

if st.button("Gerar Excel de Teste"):
    try:
        # Cria um workbook em memória
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = "Ambiente Configurado com Sucesso"
        ws['B1'] = "Davi Ladeira"
        
        # Salva o arquivo
        filename = f"{nome_arquivo}.xlsx"
        wb.save(filename)
        
        st.success(f"✅ Arquivo '{filename}' criado com sucesso!")
        st.write(f"Local: `{os.getcwd()}/{filename}`")
    except Exception as e:
        st.error(f"Erro ao criar arquivo: {e}")